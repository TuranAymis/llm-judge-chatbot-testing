"""Unit tests for the Excel reporting subsystem (specs/excel_reporting.md)."""

import importlib
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pandas as pd
import pytest


SAMPLE_ROWS = [
    {
        "Soru": "Test sorusu 1?",
        "Beklenen Cevap": "Beklenen cevap 1",
        "İlk Bot Mesajı (Hoşgeldin)": "Merhaba",
        "Chatbot'un Verdiği Gerçek Cevap": "Gerçek cevap 1",
        "AI Puanı": 5,
        "AI Durumu": "PASS",
        "AI Durumu (Ham)": "PASS",
        "AI Gerekçesi (Neden?)": "Mükemmel eşleşme",
        "Test Durumu": "PASSED",
        "Tarih": "2026-03-25 09:00:00",
    },
    {
        "Soru": "Test sorusu 2?",
        "Beklenen Cevap": "Beklenen cevap 2",
        "İlk Bot Mesajı (Hoşgeldin)": "",
        "Chatbot'un Verdiği Gerçek Cevap": "Gerçek cevap 2",
        "AI Puanı": 2,
        "AI Durumu": "FAIL",
        "AI Durumu (Ham)": "FAIL",
        "AI Gerekçesi (Neden?)": "Yetersiz bilgi",
        "Test Durumu": "FAILED",
        "Tarih": "2026-03-25 09:01:00",
    },
]


@pytest.fixture()
def reports_dir(tmp_path):
    return tmp_path / "reports"


class TestExcelCreated:
    """EC-1 & EC-5a: Normal run produces xlsx with correct row count."""

    def test_excel_created_with_data(self, reports_dir):
        from utils.excel_writer import write_excel_report

        write_excel_report(SAMPLE_ROWS, reports_dir)

        xlsx = reports_dir / "Chatbot_Test_Raporu.xlsx"
        assert xlsx.exists(), "Excel dosyası oluşturulmalı"

        wb = openpyxl.load_workbook(xlsx)
        ws = wb.active
        assert ws.max_row == 3, "1 header + 2 data row bekleniyor"
        wb.close()


class TestEmptyResults:
    """EC-5b: Empty session still produces a valid xlsx with headers only."""

    def test_empty_results_creates_headers_only(self, reports_dir):
        from utils.excel_writer import write_excel_report

        write_excel_report([], reports_dir)

        xlsx = reports_dir / "Chatbot_Test_Raporu.xlsx"
        assert xlsx.exists(), "Boş sonuçta bile dosya oluşmalı"

        wb = openpyxl.load_workbook(xlsx)
        ws = wb.active
        assert ws.max_row == 1, "Sadece header satırı bekleniyor"
        wb.close()


class TestFallbackOnPermissionError:
    """EC-4: Locked file triggers timestamped fallback."""

    def test_fallback_file_created(self, reports_dir, monkeypatch):
        from utils import excel_writer

        call_count = {"n": 0}
        _real_write = excel_writer._write_with_autowidth

        def _raise_first(df, path):
            call_count["n"] += 1
            if call_count["n"] == 1:
                raise PermissionError("file is locked")
            return _real_write(df, path)

        monkeypatch.setattr(excel_writer, "_write_with_autowidth", _raise_first)
        excel_writer.write_excel_report(SAMPLE_ROWS, reports_dir)

        fallbacks = list(reports_dir.glob("Chatbot_Test_Raporu_*.xlsx"))
        assert len(fallbacks) >= 1, "Timestamp'li yedek dosya oluşmalı"


class TestMissingOpenpyxl:
    """EC-2: openpyxl missing causes warning, not crash."""

    def test_no_crash_on_missing_openpyxl(self, reports_dir, monkeypatch, capsys):
        from utils import excel_writer

        def _raise_openpyxl(df, path):
            raise ModuleNotFoundError("No module named 'openpyxl'")

        monkeypatch.setattr(excel_writer, "_write_with_autowidth", _raise_openpyxl)
        excel_writer.write_excel_report(SAMPLE_ROWS, reports_dir)

        captured = capsys.readouterr()
        assert "openpyxl" in captured.out


class TestMissingPandas:
    """EC-2: pandas missing causes warning, not crash."""

    def test_no_crash_on_missing_pandas(self, reports_dir, capsys):
        from utils import excel_writer

        original_flag = excel_writer._HAS_PANDAS
        try:
            excel_writer._HAS_PANDAS = False
            excel_writer.write_excel_report(SAMPLE_ROWS, reports_dir)
        finally:
            excel_writer._HAS_PANDAS = original_flag

        captured = capsys.readouterr()
        assert "pandas" in captured.out
        xlsx = reports_dir / "Chatbot_Test_Raporu.xlsx"
        assert not xlsx.exists(), "pandas yokken dosya oluşmamalı"


class TestMissingOpenpyxlImportGuard:
    """openpyxl modülü yokken erken çıkış (excel_writer ön kontrol)."""

    def test_no_crash_when_openpyxl_flag_false(self, reports_dir, capsys):
        from utils import excel_writer

        orig = excel_writer._HAS_OPENPYXL
        try:
            excel_writer._HAS_OPENPYXL = False
            excel_writer.write_excel_report(SAMPLE_ROWS, reports_dir)
        finally:
            excel_writer._HAS_OPENPYXL = orig

        assert "openpyxl" in capsys.readouterr().out
        assert not (reports_dir / "Chatbot_Test_Raporu.xlsx").exists()
